"""
Excel Report Parser with SharePoint Support
This script parses Excel weekly reports and extracts information into text files.
"""

import os
#from pathlib import Path
from openpyxl import load_workbook
import requests
import msal
import io
from urllib.parse import urlparse, quote
import tempfile
from datetime import datetime
from search_sharepoint import get_sharepoint_list_items, mark_file_as_processed
from sharepoint_config import get_current_month_path, get_previous_month_path, get_specific_month_path


# SharePoint configuration
try:
    from sharepoint_config import CLIENT_ID, TENANT_ID, CLIENT_SECRET    
except ImportError:
    print("Error: sharepoint_config.py not found or has errors.")
    raise


def get_sharepoint_token(client_id=None, tenant_id=None, redirect_url=None): 
    """Get SharePoint access token using MSAL"""
    try:
        if not client_id:
            client_id = input("Enter your Azure AD application client ID: ")
        if not tenant_id:
            tenant_id = input("Enter your Azure AD tenant ID: ")
        
        # Default redirect URI for public client apps
        if not redirect_url:
            print("No redirect URI provided")
            return None
        
        # Initialize MSAL app with fixed redirect URI
        app = msal.PublicClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}"
        )
                
        scopes = ["https://graph.microsoft.com/.default"]

        # Try to get token silently first (from cache)
        accounts = app.get_accounts()
        if accounts:
            result = app.acquire_token_silent(scopes, account=accounts[0])
            if result and "access_token" in result:
                print("Using cached token")
                return result["access_token"]
        
        # Tech debt: Temporary workaround should be stored in Key Vault
        client_secret = CLIENT_SECRET
        
        app = msal.ConfidentialClientApplication(
            client_id=client_id,
            client_credential=client_secret,
            authority=f"https://login.microsoftonline.com/{tenant_id}"
        )
       
        result = app.acquire_token_for_client(scopes=scopes)
        
        if "access_token" in result:

            print("Authentication successful!")
            return result["access_token"]
        
            return None
    except Exception as e:
        print(f"Error in SharePoint authentication: {str(e)}")
        return None


def search_sharepoint_files(access_token, config=None):
    """Search for Excel files in SharePoint"""
    try:
        if config is None:
            config = SHAREPOINT_CONFIG
            
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        search_query = {
            "requests": [{
                "entityTypes": ["driveItem"],
                "query": {
                    "queryString": f'site:"{config["site"]}" path:"{config["relative_path"]}" filename:"{config["search_pattern"]}*.xlsx"'
                },
                "region": "GBR"  # Required for application permissions
            }]
        }
        
        response = requests.post(
            'https://graph.microsoft.com/v1.0/search/query',
            headers=headers,
            json=search_query
        )
        
        if response.status_code == 200:
            data = response.json()
            hits = data.get('value', [{}])[0].get('hitsContainers', [{}])[0].get('hits', [])
            return hits
        else:
            print(f"SharePoint search error: {response.status_code}")
            print(response.text)
            return []
    except Exception as e:
        print(f"Error searching SharePoint files: {str(e)}")
        return []


def download_sharepoint_file(access_token, file_url):
    """Download a file from SharePoint and return file content"""
    try:
        headers = {
            'Authorization': f'Bearer {access_token}'
        }
        
        # Convert web URL to download URL
        if 'sharepoint.com' in file_url:
            # Extract site and file path from URL
            parsed_url = urlparse(file_url)
            site_path = parsed_url.path.split('/sites/')[1].split('/')[0]
            
            # Get file info using Graph API
            site_id_url = f"https://graph.microsoft.com/v1.0/sites/jjag.sharepoint.com:/sites/{site_path}"
            site_response = requests.get(site_id_url, headers=headers)
            
            if site_response.status_code == 200:
                site_id = site_response.json()['id']
                
                # Search for the file using Graph API
                file_name = file_url.split('/')[-1]
                search_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='{file_name}')"
                search_response = requests.get(search_url, headers=headers)
                
                if search_response.status_code == 200:
                    files = search_response.json().get('value', [])
                    if files:
                        download_url = files[0]['@microsoft.graph.downloadUrl']
                        file_response = requests.get(download_url)
                        
                        if file_response.status_code == 200:
                            return file_response.content
        
        # Fallback: try direct download
        response = requests.get(file_url, headers=headers)
        if response.status_code == 200:
            return response.content
        else:
            print(f"Error downloading file: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"Error downloading SharePoint file: {str(e)}")
        return None


def process_sharepoint_workbook(access_token, file_info):
    """Process a SharePoint workbook"""
    try:
        resource = file_info.get('resource', {})
        file_name = resource.get('name', 'unknown.xlsx')
        file_url = resource.get('webUrl', '')
        
        print(f"\nProcessing SharePoint file: {file_name}")
        print(f"URL: {file_url}")
        
        # Download file content
        file_content = download_sharepoint_file(access_token, file_url)
        if not file_content:
            return False
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(file_content)
            temp_filename = temp_file.name
        
        try:
            # Process the temporary file
            success = process_workbook_content(temp_filename, file_name, access_token)
            return success
        finally:
            # Clean up temporary file
            try:
                os.unlink(temp_filename)
            except:
                pass
                
    except Exception as e:
        print(f"Error processing SharePoint workbook: {str(e)}")
        return False


def process_workbook_content_from_memory(file_content, original_name, access_token=None):
    """Process workbook content directly from memory without saving to disk"""
    try:
        # Use original filename for output
        output_name = original_name
        
        # Load the workbook from memory using BytesIO
        file_like = io.BytesIO(file_content)
        wb = load_workbook(file_like, data_only=True)
        
        return _process_workbook_data(wb, output_name, access_token)
        
    except Exception as e:
        print(f"Error processing workbook from memory {original_name}: {str(e)}")
        return False


def process_workbook_content(filename, original_name=None, access_token=None):
    """Process workbook content (extracted from original process_workbook function)"""
    try:
        # Use original filename for output if provided
        output_name = original_name if original_name else os.path.basename(filename)
        
        # Load the workbook
        wb = load_workbook(filename, data_only=True)
        
        return _process_workbook_data(wb, output_name, access_token)
        
    except Exception as e:
        print(f"Error processing workbook {filename}: {str(e)}")
        return False


def _process_workbook_data(wb, output_name, access_token=None):
    """Common workbook processing logic """
    try:
        # Get the active sheet
        sheet = wb.active
        
        # Validate that we have a sheet
        if sheet is None:
            print(f"Error: Could not access active sheet in {output_name}")
            wb.close()
            return False
        
        # First check if E33 is merged
        is_merged, target_range = is_cell_merged(sheet, "E33")
        
        if is_merged:
            print(f"Cell E33 is merged. Unmerging cells in range {target_range}")
            # Unmerge SSN range
            for merged_cell_range in list(sheet.merged_cells.ranges):
                min_col, min_row, max_col, max_row = merged_cell_range.bounds
                # Check if the merged cell range overlaps with our target range (D31:R42)
                if not (max_col < 4 or min_col > 19 or max_row < 31 or min_row > 72):
                    sheet.unmerge_cells(str(merged_cell_range))
        else:
            print("Cell E33 is not merged. No action taken.")
        
        # Create output filename
        name, ext = output_name.rsplit('.', 1) if '.' in output_name else (output_name, 'xlsx')
        fname = f"{name}.txt"

        # Build content in memory
        content_lines = []
        content_lines.append(f"Week Ending: {sheet['G7'].value}")
        content_lines.append(f"Service Provider: {sheet['G11'].value}")
        content_lines.append(f"Client: {sheet['G13'].value}")
        content_lines.append("")
        content_lines.append("Service Standard updates:")
        content_lines.append("SSN|Status|Comments")
        
        # Service standards
        for row in range(34, 43):
            if sheet[f'D{row}'].value:
                content_lines.append(f"{sheet[f'D{row}'].value}|{sheet[f'J{row}'].value}|{sheet[f'K{row}'].value}")

        # Service Risks
        content_lines.append("")
        content_lines.append("Service Risks:")
        content_lines.append("Risk No|Description|Likelihood|Impact|Mitigation")

        for row in range(45, 48):
            if sheet[f'D{row}'].value:
                content_lines.append(f"{sheet[f'D{row}'].value}|{sheet[f'E{row}'].value}|{sheet[f'H{row}'].value}|{sheet[f'J{row}'].value}|{sheet[f'K{row}'].value}")

        # Service Issues
        content_lines.append("")
        content_lines.append("Service Issues:")
        content_lines.append("Issue No|Description|Impact|Mitigation")

        for row in range(50, 53):
            if sheet[f'D{row}'].value:
                content_lines.append(f"{sheet[f'D{row}'].value}|{sheet[f'E{row}'].value}|{sheet[f'J{row}'].value}|{sheet[f'K{row}'].value}")

        # Planned Activities
        content_lines.append("")
        content_lines.append("Planned Activities:")
        content_lines.append(str(sheet['D57'].value) if sheet['D57'].value else "")

        # Client Updates
        content_lines.append("")
        content_lines.append("Client Updates:")
        content_lines.append(str(sheet['D67'].value) if sheet['D67'].value else "")

        # Join all content
        file_content = "\n".join(content_lines)
        
        # Close workbook without saving changes
        wb.close()
        
        # Upload to SharePoint if access token is provided, otherwise save locally
        if access_token:
            success = upload_text_to_sharepoint(access_token, file_content, fname)
            if success:
                return True
            else:
                print(f"⚠️ Failed to upload to SharePoint, falling back to local save: {fname}")
                # Fall back to local save
        
        # Save locally as fallback or when no access token provided
        with open(fname, 'w', encoding='utf-8') as file:
            file.write(file_content)
        print(f"📄 Successfully processed Excel data and saved locally as: {fname}")
        return True

    except Exception as e:
        print(f"Error processing workbook data: {str(e)}")
        return False

def is_cell_merged(sheet, cell_coord):
    """Check if a specific cell is part of a merged range"""
    for merged_range in sheet.merged_cells.ranges:
        if cell_coord in merged_range:
            return True, merged_range
    return False, None


def process_sharepoint_files(client_id=None, tenant_id=None, config=None, redirect_url=None):
    """Process Excel files from SharePoint using list items"""
    try:
        print("Starting SharePoint file processing using list items...")
        
        # Use config values if not provided
        if client_id is None:
            client_id = CLIENT_ID
        if tenant_id is None:
            tenant_id = TENANT_ID
        if redirect_url is None:
            redirect_url = f"msal{CLIENT_ID}://auth"
        
        # Get access token
        access_token = get_sharepoint_token(client_id, tenant_id, redirect_url)
        if not access_token:
            return False
        
        # Get SharePoint list items. 
        print("Fetching SharePoint list items...")
        # Tech debt: site and list name should be configurable
        site_name = "jjag.sharepoint.com"
        list_name = "Service Provider Uploads"
        
        list_items = get_sharepoint_list_items(access_token, site_name, list_name)
        if not list_items:
            print("No SharePoint list items found")
            return False
        
        print(f"Found {len(list_items)} items in SharePoint list")
        
        # Filter for unprocessed Excel files 
        excel_files = []
        for item in list_items:
            fields = item.get('fields', {})
            
            # Get path and filename
            path = fields.get('Path', '')
            filename = fields.get('Reportfilename', '')
            filename = f"{filename.strip()}.xlsx" 
            
            # Check if monthly report processed 
            monthly_report_processed = (fields.get('Monthlyreportprocessed'))
            if monthly_report_processed is True:
                continue  # Already processed

            manager = fields.get('manager', '')
            if manager != 'Julian Brown':
                continue  # Not the target manager 
                            
            excel_files.append({
                'filename': filename,
                'path': path,
                'full_url': f"{path.rstrip('/')}/{filename}",
                'item_id': item.get('id', ''),
                'fields': fields
            })
        
        if not excel_files:
            print("No unprocessed Excel files found in SharePoint list")
            return False
        
        print(f"Found {len(excel_files)} unprocessed Excel files to process")
        
        # Process each Excel file
        success_count = 0
        for file_info in excel_files:
            try:
                print(f"\nProcessing: {file_info['filename']}")
                
                # Download file from SharePoint
                file_content = download_sharepoint_file_from_path(access_token, file_info['path'], file_info['filename'])
                if not file_content:
                    print(f"Failed to download: {file_info['filename']}")
                    continue
                
                # Process the Excel file directly from memory
                if process_workbook_content_from_memory(file_content, file_info['filename'], access_token):
                    success_count += 1
                    print(f"Successfully processed: {file_info['filename']}")
                    
                    # Mark as processed in SharePoint
                    if mark_file_as_processed(access_token, file_info['item_id']):
                        print(f"Updated SharePoint list item - marked as processed")
                    else:
                        print(f"Warning: Failed to update SharePoint list item status")
                    
                else:
                    print(f"Failed to process Excel content: {file_info['filename']}")
                        
            except Exception as e:
                print(f"Error processing {file_info['filename']}: {str(e)}")
        
        print(f"\nSharePoint processing complete: {success_count} out of {len(excel_files)} files processed successfully")
        return success_count > 0
        
    except Exception as e:
        print(f"Error processing SharePoint files: {str(e)}")
        return False



def upload_text_to_sharepoint(access_token, file_content, filename):
    """Upload text content to SharePoint"""
    try:
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'text/plain'
        }
        
        # Get site ID first
        site_info_url = "https://graph.microsoft.com/v1.0/sites/jjag.sharepoint.com:/sites/InternalTeam:"
        site_response = requests.get(site_info_url, headers={'Authorization': f'Bearer {access_token}'})
        
        if site_response.status_code != 200:
            print(f"Failed to get site info: {site_response.status_code} - {site_response.text}")
            return False
            
        site_id = site_response.json()['id']
        
        # Try different path approaches - focusing on the new target location
        target_paths = [
            f"/MonthlyReports/{get_current_month_path()}"
        ]
        
        for target_path in target_paths:
            try:                
                # Construct file path for Graph API
                file_path = f"{target_path}/{filename}"
                
                # Clean up path - remove double slashes
                file_path = file_path.replace('//', '/')
                
                # URL encode the path
                encoded_path = quote(file_path, safe='/')
                
                # Try to upload file to SharePoint
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{encoded_path}:/content"
                
                upload_response = requests.put(upload_url, headers=headers, data=file_content.encode('utf-8'))
                
                if upload_response.status_code in [200, 201]:
                    print(f"✅ Successfully uploaded to SharePoint: {filename} at path: {target_path}")
                    return True
                else:
                    print(f"❌ Failed with path {target_path}: {upload_response.status_code} - {upload_response.text}")
                    
            except Exception as path_error:
                print(f"❌ Error with path {target_path}: {str(path_error)}")
                continue
                
    except Exception as e:
        print(f"❌ Error uploading file {filename} to SharePoint: {str(e)}")
        return False


def download_sharepoint_file_from_path(access_token, sharepoint_path, filename):
    """Download a file from SharePoint using path and filename"""
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Accept': 'application/json'
        }
        
        # Use the specified SharePoint path for downloads
        # Default to the new MonthlyReports path, but allow override via sharepoint_path parameter if needed
        if sharepoint_path and sharepoint_path.strip():
            # Clean the path - remove "Shared Documents/" if present
            clean_path = sharepoint_path
            if clean_path.startswith('Shared Documents/'):
                clean_path = clean_path[len('Shared Documents/'):]
            elif clean_path.startswith('/Shared Documents/'):
                clean_path = clean_path[len('/Shared Documents/'):]
            target_path = clean_path
        else:
            target_path = "sites/InternalTeam/Shared Documents/MonthlyReports/2025/09 - September"
        
        # Construct file path for Graph API
        if target_path:
            file_path = f"/{target_path}/{filename}"
        else:
            file_path = f"/{filename}"
        
        # Clean up path - remove double slashes
        file_path = file_path.replace('//', '/')
        
        # URL encode the path
        encoded_path = quote(file_path, safe='/')
        
        # Get site ID first
        site_info_url = "https://graph.microsoft.com/v1.0/sites/jjag.sharepoint.com:/sites/InternalTeam:"
        site_response = requests.get(site_info_url, headers=headers)
        
        if site_response.status_code != 200:
            print(f"Failed to get site info: {site_response.status_code}")
            return None
            
        site_id = site_response.json()['id']
        
        # Try to get file directly
        file_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{encoded_path}"
        
        file_response = requests.get(file_url, headers=headers)
        
        if file_response.status_code == 200:
            file_info = file_response.json()
            download_url = file_info.get('@microsoft.graph.downloadUrl')
            
            if download_url:
                # Download the actual file content
                download_response = requests.get(download_url)
                if download_response.status_code == 200:
                    print(f"Successfully downloaded: {filename}")
                    return download_response.content
        else:
            # Fallback: search for the file
            search_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/search(q='{filename}')"
            search_response = requests.get(search_url, headers=headers)
            
            if search_response.status_code == 200:
                search_results = search_response.json().get('value', [])
                
                for item in search_results:
                    if item.get('name') == filename:
                        # Get file details
                        file_id = item.get('id')
                        detail_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}"
                        detail_response = requests.get(detail_url, headers=headers)
                        
                        if detail_response.status_code == 200:
                            detail_data = detail_response.json()
                            download_url = detail_data.get('@microsoft.graph.downloadUrl')
                            
                            if download_url:
                                download_response = requests.get(download_url)
                                if download_response.status_code == 200:
                                    print(f"Successfully downloaded via search: {filename}")
                                    return download_response.content
        
        print(f"Failed to download file: {filename}")
        return None
        
    except Exception as e:
        print(f"Error downloading file {filename}: {str(e)}")
        return None


if __name__ == "__main__":
    process_sharepoint_files()




